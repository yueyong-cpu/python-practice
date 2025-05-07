# generate_report.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os

def generate_report(input_file, output_folder="."):
    # 讀取原始報價單
    df = pd.read_excel(input_file)

    # 新增「小計」欄位（單價 × 數量）
    df["小計"] = df["單價"] * df["數量"]

    # 計算總金額
    total = df["小計"].sum()

    # 在最下方新增總計欄
    total_row = pd.DataFrame({
        "品項": ["總計"],
        "單價": [""],
        "數量": [""],
        "小計": [total]
    })
    df_final = pd.concat([df, total_row], ignore_index=True)

    # 自動產生檔名，加入日期
    date_str = datetime.now().strftime("%Y%m%d")
    output_filename = f"報價總表_{date_str}.xlsx"
    output_path = os.path.join(output_folder, output_filename)

    # 寫入 Excel 並套用格式
    df_final.to_excel(output_path, index=False)

    # 開始格式化 Excel
    wb = load_workbook(output_path)
    ws = wb.active

    # 加上千分位格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row[1:]:  # 單價、數量、小計
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # 加粗並標黃總計那行（不加置中）
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    wb.save(output_path)
    print(f"✅ 報表已產出：{output_filename}")

if __name__ == "__main__":
    input_excel = "原始報價單.xlsx"
    generate_report(input_excel)
